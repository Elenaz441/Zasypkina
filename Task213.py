import csv, re, os, datetime
from typing import List, Dict, Tuple, Any
from prettytable import PrettyTable, ALL
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.workbook import Workbook
import matplotlib
import matplotlib.pyplot as plt
import numpy as np
from xlsx2html import xlsx2html
from jinja2 import Environment, FileSystemLoader
import pdfkit


class Report:
    """
    Класс для формирования отчётов и графиков.
    """
    def generate_excel(self, name_vac: str, statistic: List[Dict[str, str]]) -> None:
        """
        Генерирует Excel-таблицу.

        Args:
            name_vac (str): Название вакансии
            statistic (List[Dict[str, str]]): Статистика по вакансиям
        """
        thins = Side(border_style="thin", color="000000")
        wb = Workbook()
        sheet1 = wb['Sheet']
        sheet1.title = 'Статистика по годам'
        wb.create_sheet('Статистика по городам')
        columns1 = ['Год', 'Средняя зарплата', f'Средняя зарплата - {name_vac}',
                    'Количество вакансий', f'Количество вакансий - {name_vac}']
        for i, column in enumerate(columns1):
            sheet1.cell(row=1, column=(i + 1), value=column).font = Font(bold=True)
        for year, value in statistic[0].items():
            sheet1.append([year, value, statistic[1][year], statistic[2][year], statistic[3][year]])
        for column in sheet1.columns:
            length = max(len(str(cell.value)) for cell in column)
            sheet1.column_dimensions[column[0].column_letter].width = length + 1
            for cell in column:
                cell.border = Border(left=thins, top=thins, right=thins, bottom=thins)
                cell.alignment = Alignment(horizontal='center')
        sheet2 = wb['Статистика по городам']
        columns2 = ['Город', 'Уровень зарплат', '  ', 'Город', 'Доля вакансий']
        for i, column in enumerate(columns2):
            sheet2.cell(row=1, column=(i + 1), value=column).font = Font(bold=True)
        secondStat = list(statistic[4].items())
        for i in range(10):
            secondStat[i] += tuple(statistic[5].items())[i]
        for city1, value1, city2, value2 in secondStat:
            sheet2.append([city1, value1, '  ', city2, value2])
        for i in range(2, 12):
            sheet2[f'E{i}'].number_format = FORMAT_PERCENTAGE_00
        for column in sheet2.columns:
            length = max(len(str(cell.value)) for cell in column)
            sheet2.column_dimensions[column[0].column_letter].width = length + 1
            for cell in column:
                cell.alignment = Alignment(horizontal='center')
                if cell.value != '  ':
                    cell.border = Border(left=thins, top=thins, right=thins, bottom=thins)
        wb.save('report.xlsx')

    def generate_image(self, name_vac: str, statistic: List[Dict[str, str]]) -> None:
        """
        Генерирует файл png с графиками по вакансиям.

        Args:
            name_vac (str): Название вакансии
            statistic (List[Dict[str, str]]): Статистика по вакансиям
        """
        matplotlib.rc('font', size=8)
        width = 0.4
        fig, ((picture1, picture2), (picture3, picture4)) = plt.subplots(nrows=2, ncols=2)

        years = statistic[0].keys()
        x = np.arange(len(years))
        picture1.set_title('Уровень зарплат по годам')
        picture1.bar(x - width / 2, statistic[0].values(), width, label='средняя з/п')
        picture1.bar(x + width / 2, statistic[1].values(), width, label=f'з/п {name_vac}')
        picture1.legend(loc='upper left')
        picture1.grid(axis='y')
        picture1.set_xticks(x, years, rotation=90)

        picture2.set_title('Количество вакансий по годам')
        picture2.bar(x - width / 2, statistic[2].values(), width, label='Количество вакансий')
        picture2.bar(x + width / 2, statistic[3].values(), width, label=f'Количество вакансий {name_vac}')
        picture2.legend(loc='upper left')
        picture2.grid(axis='y')
        picture2.set_xticks(x, years, rotation=90)

        cities = list(map(lambda city: city.replace(' ', '\n').replace('-', '-\n'), list(statistic[4].keys())))
        y_pos = np.arange(len(cities))
        picture3.set_title('Уровень зарплат по городам')
        picture3.barh(y_pos, statistic[4].values(), align='center')
        picture3.invert_yaxis()  # labels read top-to-bottom
        picture3.grid(axis='x')
        picture3.set_yticks(y_pos, labels=cities, fontsize=6)

        name_labels = ['Другие'] + list(statistic[5].keys())
        city_percent = [1 - sum(list(statistic[5].values()))] + list(statistic[5].values())
        picture4.set_title('Доля вакансий по городам')
        picture4.pie(city_percent, labels=name_labels, radius=1.32, textprops={'fontsize': 6})

        fig.tight_layout()
        plt.savefig('graph.png')

    def generate_pdf(self, name_vac: str) -> None:
        """
        Генерирует файл pdf, где используются данные из прошлых методов.

        Args:
            name_vac (str): Название вакансии
        """
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")

        out1 = xlsx2html('report.xlsx', sheet='Статистика по годам')
        out1.seek(0)
        code1 = out1.read()

        out2 = xlsx2html('report.xlsx', sheet='Статистика по городам')
        out2.seek(0)
        code2 = out2.read()

        pdf_template = template.render({'name_vacancy': name_vac, 'table1': code1, 'table2': code2})

        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})


class DataSet:
    """
    Класс для формирования списка вакансий.

    Attributes:
        file_name (str): Название файла.
        vacancies_objects (List[Vacancy]): Сформированный список вакансий.

    """
    def __init__(self, file_name: str) -> None:
        """
        Инициализирует объект DataSet.

        Args:
            file_name (str): Имя файла.
        """
        self.file_name = file_name
        self.vacancies_objects = [Vacancy(vac) for vac in self.csv_filer(*self.csv_reader(file_name))]

    def clean_string(self, raw_html: str) -> str:
        """
        Очищает строку от HTML кода

        Args:
            raw_html (str): Строка, которую нужно очистить

        Returns:
            str: Очищенная строка.

        >>> DataSet('vacancies.csv').clean_string('<p>Группа компаний «МИАКОМ»</p>')
        'Группа компаний «МИАКОМ»'
        """
        result = re.sub("<.*?>", '', raw_html)
        return result if '\n' in raw_html else " ".join(result.split())

    def csv_reader(self, file_name: str) -> Tuple[List[str], List[List[str]]]:
        """
        Считывает данные из файла

        Args:
            file_name (str): Название файла.

        Returns:
            Tuple[List[str], List[List[str]]]: Название колонок и соответствующие им данные по каждой вакансии.
        """
        reader = csv.reader(open(file_name, encoding='utf_8_sig'))
        data_base = [line for line in reader]
        return data_base[0], data_base[1:]

    def csv_filer(self, list_naming: List[str], reader: List[List[str]]) -> List[Dict[str, str]]:
        """
        Преобразует данные в список словарей, где словарь содержит информацию об одной вакансии.

        Args:
            list_naming (List[str]): Поля вакансии
            reader (List[List[str]]): Данные из файла

        Returns:
            List[Dict[str, str]]: Список словарей.
        """
        new_vacans_list = list(filter(lambda vac: (len(vac) == len(list_naming) and vac.count('') == 0), reader))
        return [dict(zip(list_naming, map(self.clean_string, vac))) for vac in new_vacans_list]


class Vacancy:
    """
    Класс для представления вакансии
    
    Attributes:
        name (string): Название вакансии
        description (str): Описание вакансии
        key_skills (List[str]): Ключевые навыки для вакансии
        experience_id (str): Требуемый опят для вакансии
        premium (str): Атрибут, отвечающий за премиальность вакансии
        employer_name (str): Название компании, где есть вакансия
        salary (Salary): Информация о зарплате
        area_name (str): Название города
        published_at (str): Дата публикации вакансии
    """
    def __init__(self, dict_vac: Dict[str, str]):
        """
        Инициализирует объект Vacancy, проверяя наличие некоторых полей для вакансии

        Args: dict_vac (Dict[str, str]): Словарь, хранящий информацию о вакансии. Ключи - это названия полей,
        значения - информация о вакансии по соответствующему полю.
        """
        self.name = dict_vac['name']
        self.description = 'Нет данных' if 'description' not in dict_vac.keys() else dict_vac['description']
        self.key_skills = 'Нет данных' if 'key_skills' not in dict_vac.keys() else dict_vac['key_skills'].split('\n')
        self.experience_id = 'Нет данных' if 'experience_id' not in dict_vac.keys() else dict_vac['experience_id']
        self.premium = 'Нет данных' if 'premium' not in dict_vac.keys() else dict_vac['premium']
        self.employer_name = 'Нет данных' if 'employer_name' not in dict_vac.keys() else dict_vac['employer_name']
        salary_gross = 'Нет данных' if 'salary_gross' not in dict_vac.keys() else dict_vac['salary_gross']
        self.salary = Salary(dict_vac['salary_from'], dict_vac['salary_to'], salary_gross, dict_vac['salary_currency'])
        self.area_name = dict_vac['area_name']
        self.published_at = dict_vac['published_at']


class Salary:
    """
    Класс для представления зарплаты.

    Attributes:
        salary_from (str): Нижняя граница зарплаты
        salary_to (str): Верхняя граница зарплаты
        salary_gross (str): Наличие налогов
        salary_currency (str): Валюта оклада

    """
    def __init__(self, salary_from, salary_to, salary_gross, salary_currency):
        """
        Инициализирует объект Salary
        Args:
            salary_from (str): Нижняя граница зарплаты
            salary_to (str): Верхняя граница зарплаты
            salary_gross (str): Наличие налогов
            salary_currency (str): Валюта оклада
        """
        self.salary_from = salary_from
        self.salary_to = salary_to
        self.salary_gross = salary_gross
        self.salary_currency = salary_currency

    def to_RUB(self, salary: float) -> float:
        """
        Вычисляет зарплату в рублях, при помощи словаря - currency_to_rub.

        Args:
            salary (float): Зарплата в другой валюте.

        Returns:
            float: Зарплата в рублях.

        >>> Salary('10', '1000', 'true', 'EUR').to_RUB(1000.0)
        59900.0
        >>> Salary('10', '1000', 'true', 'RUR').to_RUB(1000)
        1000.0
        >>> Salary('10', '1000', 'true', 'QWE').to_RUB(1000.0)
        Traceback (most recent call last):
        ...
        KeyError: 'QWE'
        """
        return float(salary * currency_to_rub[self.salary_currency])


class InputConnect:
    """
    Формирование таблицы PrettyTable с удобным отображением информации о вакансии.

    Attributes:
        filter_param (str or List[str]): Параметр фильтрации
        sort_param (str): Параметр сортировки
        reversed_sort (str or bool): Параметр обратной сортировки
        interval (List[int]): Промежуток выводимых колонок
        columns (str or List[str]): Название выводимых колонок
    """
    def __init__(self, filter_param, sort_param, reversed_sort, interval, columns):
        """
        Инициализирует объект InputConnect

        Args:
            filter_param (str): Параметр фильтрации
            sort_param (str): Параметр сортировки
            reversed_sort (str or bool): Параметр обратной сортировки
            interval (List[int]): Промежуток выводимых колонок
            columns (str or List[str]): Название выводимых колонок
        """
        self.filter_param = filter_param
        self.sort_param = sort_param
        self.reversed_sort = reversed_sort
        self.interval = interval
        self.columns = columns

    def check_parameters(self) -> None:
        """
        Проверка параметров на корректность ввода.
        """
        if ': ' not in self.filter_param and self.filter_param != '':
            exit_from_file('Формат ввода некорректен')
        self.filter_param = self.filter_param.split(': ')
        if len(self.filter_param) == 2 and self.filter_param[0] not in list(translation.values()):
            exit_from_file('Параметр поиска некорректен')
        if self.sort_param != '' and self.sort_param not in list(translation.values()):
            exit_from_file('Параметр сортировки некорректен')
        if self.reversed_sort not in ['Да', 'Нет', '']:
            exit_from_file('Порядок сортировки задан некорректно')
        self.reversed_sort = (self.reversed_sort == 'Да')
        if len(self.columns) != 0:
            self.columns = self.columns.split(', ')
            self.columns.insert(0, '№')

    def formatter(self, vacancy: Vacancy) -> List[Any]:
        """
        Осуществляет форматирование необходимых полей.

        Args:
            vacancy (Vacancy): Вакансия.

        Returns:
            List[Any]: Список отформатированных полей.
        """
        def change_salary(salary: Salary) -> str:
            """
            Форматирует зарплату к нужному формату.

            Args:
                salary (Salary): Информация о зарплате.

            Returns:
                str: Отформатированная информация о зарплате.
            """
            salary_from = int(float(salary.salary_from))
            salary_to = int(float(salary.salary_to))
            if salary_from >= 1000:
                salary_from = f'{salary_from // 1000} {str(salary_from)[-3:]}'
            if salary_to >= 1000:
                salary_to = f'{salary_to // 1000} {str(salary_to)[-3:]}'
            info_gross = 'Без вычета налогов' if translation[salary.salary_gross] == 'Да' else 'С вычетом налогов'
            result_salary = f'{salary_from} - {salary_to} ({translation[salary.salary_currency]}) ({info_gross})'
            return result_salary

        def change_date(date_vac: str) -> str:
            """
            Форматирует дату публикации к нужному формату.

            Args:
                date_vac (str): Дата публикации.

            Returns:
                str: Отформатированная дата публикации.
            """
            return datetime.datetime.strptime(date_vac, '%Y-%m-%dT%H:%M:%S%z').strftime('%d.%m.%Y')

        return [vacancy.name, vacancy.description, '\n'.join(vacancy.key_skills), translation[vacancy.experience_id],
                translation[vacancy.premium], vacancy.employer_name, change_salary(vacancy.salary), vacancy.area_name,
                change_date(vacancy.published_at)]

    def data_filter(self, list_vacancies: List[Vacancy], parameter: List[str]) -> List[Vacancy]:
        """
        Фильтрует список вакансий по введённым параметрам.

        Args:
            list_vacancies (List[Vacancy]): Список вакансий.
            parameter (List[str]): Параметры фильтрации.

        Returns:
            List[Vacancy]: Список отфильтрованных вакансий.
        """
        if parameter[0] == 'Навыки':
            parameter[1] = parameter[1].split(', ')
        if parameter[0] == 'Оклад':
            list_vacancies = list(
                filter(lambda vac: int(vac.salary.salary_from) <= int(parameter[1]) <= int(vac.salary.salary_to),
                       list_vacancies))
        elif parameter[0] == 'Навыки':
            list_vacancies = list(
                filter(lambda vac: all(item in vac.key_skills for item in parameter[1]), list_vacancies))
        elif parameter[0] == 'Опыт работы' or parameter[0] == 'Премиум-вакансия':
            list_vacancies = list(
                filter(lambda vac: parameter[1] == translation[vac.__getattribute__(reverse_translation[parameter[0]])],
                       list_vacancies))
        elif parameter[0] == 'Идентификатор валюты оклада':
            list_vacancies = list(
                filter(lambda vac: parameter[1] == translation[vac.salary.salary_currency], list_vacancies))
        elif parameter[0] == 'Дата публикации вакансии':
            list_vacancies = list(filter(lambda vac: parameter[1] == datetime.datetime.strptime(vac.published_at,
                                                '%Y-%m-%dT%H:%M:%S%z').strftime('%d.%m.%Y'), list_vacancies))
        else:
            list_vacancies = list(
                filter(lambda vac: parameter[1] == vac.__getattribute__(reverse_translation[parameter[0]]),
                       list_vacancies))
        return list_vacancies

    def data_sort(self, list_vacancies: List[Vacancy], param: str, is_reverse: bool) -> List[Vacancy]:
        """
        Сортирует список вакансий по введённым параметрам.

        Args:
            list_vacancies (List[Vacancy]): Список вакансий.
            param (str): Параметр сортировки.
            is_reverse (bool): Параметр обратной сортировки.

        Returns:
            List[Vacancy]: Список отсортированных вакансий.
        """
        if param == 'Навыки':
            list_vacancies.sort(key=lambda vac: len(vac.key_skills), reverse=is_reverse)
        elif param == 'Оклад':
            list_vacancies.sort(
                key=lambda vac: vac.salary.to_RUB(float(vac.salary.salary_from) + float(vac.salary.salary_to)) / 2,
                reverse=is_reverse)
        elif param == 'Дата публикации вакансии':
            list_vacancies.sort(key=lambda vac: datetime.datetime.strptime(vac.published_at, '%Y-%m-%dT%H:%M:%S%z'),
                                reverse=is_reverse)
        elif param == 'Опыт работы':
            list_vacancies.sort(key=lambda vac: rang_experience_id[vac.experience_id], reverse=is_reverse)
        else:
            list_vacancies.sort(key=lambda vac: vac.__getattribute__(reverse_translation[param]), reverse=is_reverse)
        return list_vacancies

    def print_vacancies(self, list_vacancies: List[Vacancy]) -> None:
        """
        Выводит информацию о вакансии в таблицу PrettyTable

        Args:
            list_vacancies (List[Vacancy]): Список вакансий.
        """
        self.interval.append(len(list_vacancies) + 1)
        list_vacancies = list_vacancies if len(self.filter_param) != 2 else self.data_filter(list_vacancies, self.filter_param)
        list_vacancies = list_vacancies if len(list_vacancies) != 0 else 'Ничего не найдено'
        if type(list_vacancies) is str:
            print(list_vacancies)
            return
        list_vacancies = list_vacancies if len(self.sort_param) == 0 else self.data_sort(list_vacancies, self.sort_param, self.reversed_sort)
        table_header = list(reverse_translation.keys())[:-1]
        table_header.insert(0, '№')
        vacans_table = PrettyTable(table_header)
        vacans_table.hrules = ALL
        for i in range(len(list_vacancies)):
            vac = self.formatter(list_vacancies[i])
            vac = list(map(lambda i: f'{i[:100]}...' if len(i) > 100 else i, vac))
            vac.insert(0, i + 1)
            vacans_table.add_row(vac)
        vacans_table.align = 'l'
        vacans_table.max_width = 20
        if len(self.interval) > 1 and len(self.columns) >= 2:
            vacans_table = vacans_table.get_string(start=self.interval[0] - 1, end=self.interval[1] - 1, fields=self.columns)
        elif len(self.interval) > 1:
            vacans_table = vacans_table.get_string(start=self.interval[0] - 1, end=self.interval[1] - 1)
        elif len(self.columns) >= 2:
            vacans_table = vacans_table.get_string(fields=self.columns)
        print(vacans_table)


def get_salary_level(list_vacancies: List[Vacancy], field: str, name_vacancy: str = '') -> Dict[str, str]:
    """
    Формирует статистики, связанные с зарплатами

    Args:
        list_vacancies (List[Vacancy]): Список вакансий
        field (str): Поле вакансии
        name_vacancy (str): Название вакансии (если его ввели)

    Returns:
        Dict[str, str]: Статистика связанная с зарплатой
    """
    result = {}
    for vac in list_vacancies:
        result[vac.__getattribute__(field)] = [] if vac.__getattribute__(field) not in result.keys() else result[vac.__getattribute__(field)]
    list_vacancies = list_vacancies if name_vacancy == '' else list(filter(lambda vac: name_vacancy in vac.name, list_vacancies))
    for vac in list_vacancies:
        result[vac.__getattribute__(field)].append(vac.salary.to_RUB(float(vac.salary.salary_from) + float(vac.salary.salary_to)) / 2)
    for key in result.keys():
        result[key] = 0 if len(result[key]) == 0 else int(sum(result[key]) // len(result[key]))
    return result


def get_count_vacancies(list_vacancies: List[Vacancy], field: str, name_vacancy: str = '') -> Dict[str, str]:
    """
    Формирует статистики, связанные с количеством вакансий

    Args:
        list_vacancies (List[Vacancy]): Список вакансий
        field (str): Поле вакансии
        name_vacancy (str): Название вакансии (если его ввели)

    Returns:
        Dict[str, str]: Статистика, связанная с количеством вакансий
    """
    result = {}
    for vac in list_vacancies:
        result[vac.__getattribute__(field)] = 0 if vac.__getattribute__(field) not in result.keys() else result[vac.__getattribute__(field)]
    list_vacancies = list_vacancies if name_vacancy == '' else list(filter(lambda vac: name_vacancy in vac.name, list_vacancies))
    for vac in list_vacancies:
        result[vac.__getattribute__(field)] += 1
    if field == 'area_name':
        for key in result.keys():
            result[key] = round(result[key] / len(data.vacancies_objects), 4)
    return result


currency_to_rub = {"AZN": 35.68,
                   "BYR": 23.91,
                   "EUR": 59.90,
                   "GEL": 21.74,
                   "KGS": 0.76,
                   "KZT": 0.13,
                   "RUR": 1,
                   "UAH": 1.64,
                   "USD": 60.66,
                   "UZS": 0.0055, }

translation = {"name": "Название",
               "description": "Описание",
               "key_skills": "Навыки",
               "experience_id": "Опыт работы",
               "premium": "Премиум-вакансия",
               "employer_name": "Компания",
               "salary_from": "Нижняя граница вилки оклада",
               "salary_to": "Верхняя граница вилки оклада",
               "salary_gross": "Оклад указан до вычета налогов",
               "salary_currency": "Идентификатор валюты оклада",
               "area_name": "Название региона",
               "published_at": "Дата публикации вакансии",
               "Оклад": "Оклад",
               "Нет данных": "Нет данных",
               "True": "Да",
               "TRUE": "Да",
               "False": "Нет",
               "FALSE": "Нет",
               "noExperience": "Нет опыта",
               "between1And3": "От 1 года до 3 лет",
               "between3And6": "От 3 до 6 лет",
               "moreThan6": "Более 6 лет",
               "AZN": "Манаты",
               "BYR": "Белорусские рубли",
               "EUR": "Евро",
               "GEL": "Грузинский лари",
               "KGS": "Киргизский сом",
               "KZT": "Тенге",
               "RUR": "Рубли",
               "UAH": "Гривны",
               "USD": "Доллары",
               "UZS": "Узбекский сум"}

reverse_translation = {"Название": "name",
                       "Описание": "description",
                       "Навыки": "key_skills",
                       "Опыт работы": "experience_id",
                       "Премиум-вакансия": "premium",
                       "Компания": "employer_name",
                       "Оклад": "Оклад",
                       "Название региона": "area_name",
                       "Дата публикации вакансии": "published_at",
                       "Идентификатор валюты оклада": "salary_currency"}

rang_experience_id = {"noExperience": 0,
                      "between1And3": 1,
                      "between3And6": 2,
                      "moreThan6": 3}


def change_data(date_vac) -> str:
    """
    Форматирует дату публикации к нужному формату.

    Args:
        date_vac (str): Дата публикации.

    Returns:
        str: Отформатированная дата публикации.
    """
    return datetime.datetime.strptime(date_vac, '%Y-%m-%dT%H:%M:%S%z').strftime('%Y')


def exit_from_file(message: str):
    """
    Метод для выхода из программы.

    Args:
        message (str): Сообщение при выходе из программы
    """
    print(message)
    exit()


def get_statistic(result_list: Dict[Any, Any], index: int, is_reversed: bool = False, slice: int = 0) -> Dict[Any, Any]:
    """
    Приводит статистику к нужному виду (чтобы года шли )

    Args:
        result_list (Dict[Any, Any]): Словарь со статистикой
        index (int): Индекс
        is_reversed (bool): Отвечает за обратную сортировку
        slice (int): Срез для статистики

    Returns:
        Dict[Any, Any]:
    """
    slice = len(result_list) if slice == 0 else slice
    return dict(sorted(result_list, key=lambda x: x[index], reverse=is_reversed)[:slice])


type_output = input('Введите данные для печати: ')
file_name = input('Введите название файла: ')
if os.stat(file_name).st_size == 0:
    exit_from_file('Пустой файл')
data = DataSet(file_name)
if len(data.vacancies_objects) == 0:
    exit_from_file('Нет данных')
if type_output == 'Статистика':
    vacancy_name = input('Введите название профессии: ')
    for vac in data.vacancies_objects:
        vac.published_at = change_data(vac.published_at)
    dict_cities = {}
    for vac in data.vacancies_objects:
        if vac.area_name not in dict_cities.keys():
            dict_cities[vac.area_name] = 0
        dict_cities[vac.area_name] += 1
    needed_vacancies_objects = list(filter(lambda vac: int(len(data.vacancies_objects) * 0.01) <= dict_cities[vac.area_name], data.vacancies_objects))
    rp = Report()
    list_statistic = [get_statistic(get_salary_level(data.vacancies_objects, 'published_at').items(), 0),
                      get_statistic(get_salary_level(data.vacancies_objects, 'published_at', vacancy_name).items(), 0),
                      get_statistic(get_count_vacancies(data.vacancies_objects, 'published_at').items(), 0),
                      get_statistic(get_count_vacancies(data.vacancies_objects, 'published_at', vacancy_name).items(), 0),
                      get_statistic(get_salary_level(needed_vacancies_objects, 'area_name').items(), 1, True, 10),
                      get_statistic(get_count_vacancies(needed_vacancies_objects, 'area_name').items(), 1, True, 10)]
    rp.generate_excel(vacancy_name, list_statistic)
    rp.generate_image(vacancy_name, list_statistic)
    rp.generate_pdf(vacancy_name)
elif type_output == 'Вакансии':
    parameter = input('Введите параметр фильтрации: ')
    sorting_param = input('Введите параметр сортировки: ')
    is_reversed_sort = input('Обратный порядок сортировки (Да / Нет): ')
    interval = list(map(int, input('Введите диапазон вывода: ').split()))
    columns = input('Введите требуемые столбцы: ')
    outer = InputConnect(parameter, sorting_param, is_reversed_sort, interval, columns)
    outer.check_parameters()
    outer.print_vacancies(data.vacancies_objects)