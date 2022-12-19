import csv, re, os
from typing import List, Dict, Tuple, Any
from openpyxl.styles import Font, Border, Side, Alignment
from openpyxl.styles.numbers import FORMAT_PERCENTAGE_00
from openpyxl.workbook import Workbook
import matplotlib
import matplotlib.pyplot as plt
import numpy as np
from xlsx2html import xlsx2html
from jinja2 import Environment, FileSystemLoader
import pdfkit


class Report:
    """
    Класс для формирования отчётов и графиков.
    """

    def generate_excel(self, name_vac: str, statistic: List[Dict[str, str]]) -> None:
        """
        Генерирует Excel-таблицу.

        Args:
            name_vac (str): Название вакансии
            statistic (List[Dict[str, str]]): Статистика по вакансиям
        """
        thins = Side(border_style="thin", color="000000")
        wb = Workbook()
        sheet1 = wb['Sheet']
        sheet1.title = 'Статистика по годам'
        wb.create_sheet('Статистика по городам')
        columns1 = ['Год', 'Средняя зарплата', f'Средняя зарплата - {name_vac}',
                    'Количество вакансий', f'Количество вакансий - {name_vac}']
        for i, column in enumerate(columns1):
            sheet1.cell(row=1, column=(i + 1), value=column).font = Font(bold=True)
        for year, value in statistic[0].items():
            sheet1.append([year, value, statistic[1][year], statistic[2][year], statistic[3][year]])
        for column in sheet1.columns:
            length = max(len(str(cell.value)) for cell in column)
            sheet1.column_dimensions[column[0].column_letter].width = length + 1
            for cell in column:
                cell.border = Border(left=thins, top=thins, right=thins, bottom=thins)
                cell.alignment = Alignment(horizontal='center')
        sheet2 = wb['Статистика по городам']
        columns2 = ['Город', 'Уровень зарплат', '  ', 'Город', 'Доля вакансий']
        for i, column in enumerate(columns2):
            sheet2.cell(row=1, column=(i + 1), value=column).font = Font(bold=True)
        secondStat = list(statistic[4].items())
        for i in range(10):
            secondStat[i] += tuple(statistic[5].items())[i]
        for city1, value1, city2, value2 in secondStat:
            sheet2.append([city1, value1, '  ', city2, value2])
        for i in range(2, 12):
            sheet2[f'E{i}'].number_format = FORMAT_PERCENTAGE_00
        for column in sheet2.columns:
            length = max(len(str(cell.value)) for cell in column)
            sheet2.column_dimensions[column[0].column_letter].width = length + 1
            for cell in column:
                cell.alignment = Alignment(horizontal='center')
                if cell.value != '  ':
                    cell.border = Border(left=thins, top=thins, right=thins, bottom=thins)
        wb.save('report.xlsx')

    def generate_image(self, name_vac: str, statistic: List[Dict[str, str]]) -> None:
        """
        Генерирует файл png с графиками по вакансиям.

        Args:
            name_vac (str): Название вакансии
            statistic (List[Dict[str, str]]): Статистика по вакансиям
        """
        matplotlib.rc('font', size=8)
        width = 0.4
        fig, ((picture1, picture2), (picture3, picture4)) = plt.subplots(nrows=2, ncols=2)

        years = statistic[0].keys()
        x = np.arange(len(years))
        picture1.set_title('Уровень зарплат по годам')
        picture1.bar(x - width / 2, statistic[0].values(), width, label='средняя з/п')
        picture1.bar(x + width / 2, statistic[1].values(), width, label=f'з/п {name_vac}')
        picture1.legend(loc='upper left')
        picture1.grid(axis='y')
        picture1.set_xticks(x, years, rotation=90)

        picture2.set_title('Количество вакансий по годам')
        picture2.bar(x - width / 2, statistic[2].values(), width, label='Количество вакансий')
        picture2.bar(x + width / 2, statistic[3].values(), width, label=f'Количество вакансий {name_vac}')
        picture2.legend(loc='upper left')
        picture2.grid(axis='y')
        picture2.set_xticks(x, years, rotation=90)

        cities = list(map(lambda city: city.replace(' ', '\n').replace('-', '-\n'), list(statistic[4].keys())))
        y_pos = np.arange(len(cities))
        picture3.set_title('Уровень зарплат по городам')
        picture3.barh(y_pos, statistic[4].values(), align='center')
        picture3.invert_yaxis()  # labels read top-to-bottom
        picture3.grid(axis='x')
        picture3.set_yticks(y_pos, labels=cities, fontsize=6)

        name_labels = ['Другие'] + list(statistic[5].keys())
        city_percent = [1 - sum(list(statistic[5].values()))] + list(statistic[5].values())
        picture4.set_title('Доля вакансий по городам')
        picture4.pie(city_percent, labels=name_labels, radius=1.32, textprops={'fontsize': 6})

        fig.tight_layout()
        plt.savefig('graph.png')

    def generate_pdf(self, name_vac: str) -> None:
        """
        Генерирует файл pdf, где используются данные из прошлых методов.

        Args:
            name_vac (str): Название вакансии
        """
        env = Environment(loader=FileSystemLoader('.'))
        template = env.get_template("pdf_template.html")

        out1 = xlsx2html('report.xlsx', sheet='Статистика по годам')
        out1.seek(0)
        code1 = out1.read()

        out2 = xlsx2html('report.xlsx', sheet='Статистика по городам')
        out2.seek(0)
        code2 = out2.read()

        pdf_template = template.render({'name_vacancy': name_vac, 'table1': code1, 'table2': code2})

        config = pdfkit.configuration(wkhtmltopdf=r'C:\Program Files\wkhtmltopdf\bin\wkhtmltopdf.exe')
        pdfkit.from_string(pdf_template, 'report.pdf', configuration=config, options={"enable-local-file-access": ""})


def get_salary_level(list_vacancies: List[Any], field: str, name_vacancy: str = '') -> Dict[str, str]:
    """
    Формирует статистики, связанные с зарплатами

    Args:
        list_vacancies (List[Vacancy]): Список вакансий
        field (str): Поле вакансии
        name_vacancy (str): Название вакансии (если его ввели)

    Returns:
        Dict[str, str]: Статистика связанная с зарплатой
    """
    result = {}
    for vac in list_vacancies:
        result[vac.__getattribute__(field)] = [] if vac.__getattribute__(field) not in result.keys() else result[
            vac.__getattribute__(field)]
    list_vacancies = list_vacancies if name_vacancy == '' else list(
        filter(lambda vac: name_vacancy in vac.name, list_vacancies))
    for vac in list_vacancies:
        result[vac.__getattribute__(field)].append(
            vac.salary.to_RUB(float(vac.salary.salary_from) + float(vac.salary.salary_to)) / 2)
    for key in result.keys():
        result[key] = 0 if len(result[key]) == 0 else int(sum(result[key]) // len(result[key]))
    return result


def get_count_vacancies(list_vacancies: List[Any], field: str, data: Any, name_vacancy: str = '') -> Dict[str, str]:
    """
    Формирует статистики, связанные с количеством вакансий

    Args:
        list_vacancies (List[Vacancy]): Список вакансий
        field (str): Поле вакансии
        data(DataSet): Данные из файла
        name_vacancy (str): Название вакансии (если его ввели)

    Returns:
        Dict[str, str]: Статистика, связанная с количеством вакансий
    """
    result = {}
    for vac in list_vacancies:
        result[vac.__getattribute__(field)] = 0 if vac.__getattribute__(field) not in result.keys() else result[
            vac.__getattribute__(field)]
    list_vacancies = list_vacancies if name_vacancy == '' else list(
        filter(lambda vac: name_vacancy in vac.name, list_vacancies))
    for vac in list_vacancies:
        result[vac.__getattribute__(field)] += 1
    if field == 'area_name':
        for key in result.keys():
            result[key] = round(result[key] / len(data.vacancies_objects), 4)
    return result


def print_statistic(result_list, index, message, is_reversed=False, slice=0):
    """
    Приводит статистику к нужному виду (чтобы года шли по порядку) и печатает её

    Args:
        result_list (Dict[Any, Any]): Словарь со статистикой
        index (int): Индекс
        is_reversed (bool): Отвечает за обратную сортировку
        slice (int): Срез для статистики

    Returns:
        Dict[Any, Any]: Преобразованный словарь со статистикой
    """
    slice = len(result_list) if slice == 0 else slice
    statistic = dict(sorted(result_list, key=lambda x: x[index], reverse=is_reversed)[:slice])
    print(message + str(statistic))
    return statistic


def get_statistic(result_list: Dict[Any, Any], index: int, is_reversed: bool = False, slice: int = 0) -> Dict[Any, Any]:
    """
    Приводит статистику к нужному виду (чтобы года шли по порядку)

    Args:
        result_list (Dict[Any, Any]): Словарь со статистикой
        index (int): Индекс
        is_reversed (bool): Отвечает за обратную сортировку
        slice (int): Срез для статистики

    Returns:
        Dict[Any, Any]: Преобразованный словарь со статистикой
    """
    slice = len(result_list) if slice == 0 else slice
    return dict(sorted(result_list, key=lambda x: x[index], reverse=is_reversed)[:slice])
